# ========================================================================== #
#               -----   xlsx_Mngr.py   -----                                 #
#              class  for  xlsx file managiging                              #
#  for more informations see Data_Organization.txt                           #
# ========================================================================== #

from Data_Classes.Codes_DB import *
from openpyxl import load_workbook
from Common.Common_Functions import *
from datetime import datetime

class Xlsx_Manager(Codes_db):
    def __init__(self):
        super().__init__()

        # one row on xlsx file   ---------------------------------------------------------------
        self.Contab = None
        self.Valuta = None
        self.Des1   = None
        self.Accr   = None
        self.Addeb  = None
        self.Des2   = None
        #                  nRow Contab Valuta TR_Desc Accred Addeb TRcode
        #                #   x    1       2      x       4     5     6
        self.ItemToCheck = [99,   1,      2,    99,      4,    5,    6]
        self.strList     = ['Row.    ',  'Contab: ', 'Valuta: ', 'Descr:  ', 'Accr:   ', 'Addeb:  ', '']

        # ------------------------             A      B      C      D      E     F  ---------------
        self._Xlsx_Rows_From_Sheet   = []  # Contab Valuta Descr1 Accred Addeb Descr2
        self._Xlsx_Rows_Desc_Compact = []
        self._Xlsx_Rows_NOK_List     = []
        #
        self._With_Code_Tree_List    = []  # nRow Contabile Valuta TRdesc Accred Addeb TRcode
        self._Wihtout_Code_Tree_List = []  # nRow Valuta Descr
        # ------------------------
        self._Tot_Rows        = 0
        self._Tot_NOK         = 0
        self._Tot_OK          = 0
        self._TotWith_Code    = 0
        self._TotWihtout_Code = 0
        self._iYear_List      = []

        # -------   _tAtt : temporary attributes  that will be coied on _Att  if all OK  -------------
        self._tXLSX_Rows_From_Sheet    = []
        self._tXLSX_Rows_Desc_Compact  = []
        self._tXlsx_Rows_NOK_List      = []
        #
        self._tWith_Code_Tree_List     = []  # nRow Contabile Valuta TRdesc Accred Addeb TRcode
        self._tWihtout_Code_Tree_List  = []
        # ------------------------
        self._tTot_Rows     = 0
        self._tTot_OK       = 0
        self._tTot_NOK      = 0
        self._tTotWith_Code    = 0
        self._tTotWihtout_Code = 0
        self._tiYear_List   = []
        # -----------------------------------------------------------------------------------------
        self._tXlsx_Conto    = None  # these attributs are not saved on Selections
        self._tXlsx_Year     = None  # calculated in _Set_Xlsx_Conto_Year_Month()
        self._tXlsx_Month    = None  #    """      "     """      """

    # ----------------------------------------------------------------------------------- #
    #            ----------------      public   methods   -----------------               #
    # ----------------------------------------------------------------------------------- #
    def Get_Total_Rows(self):
        # Ix_Tot_OK, Ix_Tot_WithCode, Ix_Tot_Without_Code
        return [self._Tot_OK, self._TotWith_Code, self._TotWihtout_Code]

    def Get_WithCodeList(self):
        return self._With_Code_Tree_List

    def Get_WithoutCodeList(self):
        return self._Wihtout_Code_Tree_List

    def Get_Xlsx_Rows_From_Sheet(self):
        return self._Xlsx_Rows_From_Sheet

    def Clear_Xlsx_Conto_Year_Month(self):
        self._tXlsx_Conto = None
        self._tXlsx_Year  = None
        self._tXlsx_Month = None

    def _Set_Xlsx_Conto_Year_Month(self, Filename):
        FullFilename = Filename  # self.Get_Selections_Member(Ix_Xlsx_File)
        if FullFilename != UNKNOWN:
            filename = Get_File_Name(FullFilename)
            self._tXlsx_Conto = filename[0:5]
            self._tXlsx_Year = int(filename[6:10])
            self._tXlsx_Month = int(filename[11:13])
        else:
            self._tXlsx_Conto = None
            self._tXlsx_Year  = None
            self._tXlsx_Month = None

    # --------------------------------------------------------------------------------------------
    def _Init_Xlsx_Data(self):
        self._tXLSX_Rows_From_Sheet   = []
        self._tXLSX_Rows_Desc_Compact = []
        self._tXlsx_Rows_NOK_List     = []
        #
        self._tWith_Code_Tree_List    = []  # nRow Contabile Valuta TRdesc Accred Addeb TRcode
        self._tWihtout_Code_Tree_List = []
        # ------------------------
        self._tTot_Rows     = 0
        self._tTot_OK       = 0
        self._tTot_NOK      = 0
        self._tTotWith_Code = 0
        self._tTotWihtout_Code = 0
        self._tiYear_List = []
        # -----------------------------------------------------------------------------------------
        self._tXlsx_Conto    = None  # or on selecting new file  FIDEU_2024_01.xlsx
        self._tXlsx_Year     = None  # they are  calculated on startup
        self._tXlsx_Month    = None

    # --------------------------------------------------------------------------------------------
    def _Save_Xlsx_Data(self):
        self._Xlsx_Rows_From_Sheet   = self._tXLSX_Rows_From_Sheet
        self._Xlsx_Rows_Desc_Compact = self._tXLSX_Rows_Desc_Compact
        self._Xlsx_Rows_NOK_List     = self._tXlsx_Rows_NOK_List
        #
        self._With_Code_Tree_List    = self._tWith_Code_Tree_List
        self._Wihtout_Code_Tree_List = self._tWihtout_Code_Tree_List
        # ------------------------
        self._Tot_Rows     = self._tTot_Rows
        self._Tot_OK       = self._tTot_OK
        self._Tot_NOK      = self._tTot_NOK
        self._TotWith_Code = self._tTot_WithCode
        self._TotWihtout_Code = self._tTot_WithoutCode
        self._iYear_List    = self._tiYear_List

        self._Xlsx_Conto = self._tXlsx_Conto
        self._Xlsx_Year  = self._tXlsx_Year
        self._Xlsx_Month = self._tXlsx_Month

    # -----------------------------------------------------------------------------------------
    def Load_Xlsx_Lists(self, Filename):
        File_Name    = Filename
        if Filename == ON_SELECTIONS:
            File_Name = self._Xlsx_Filename                     # set filename
        if File_Name  == UNKNOWN:
            return 'Xlsx filename unknown'          # return UNKNOWN

        self._Init_Xlsx_Data()
        # to adjust the rows values for different Conto (FIDEU_2024_09.xlsx)
        self._Set_Xlsx_Conto_Year_Month(File_Name)

        Result = self._Load_xlsx_Rows_From_Sheet(File_Name)      # Load xlsx Rows
        if Result != OK:
            return Result                           # return  'Error on loading' or 'No rows'

        else:
            Result = self._Create_Xlsx_Lists()                  # create xlsx lists
            if Result == OK:
                self._Save_Xlsx_Data()
                self._Files_Loaded[Ix_Xlsx_Lists_Loaded] = True
                return OK
            else:
                return Result                       # return   'Multi matching'

    # -------------------------------------  Get rows from sheet ----------------------------
    def _Load_xlsx_Rows_From_Sheet(self, Filename):
        self._Get_Work_Sheet_Rows(Filename)  # set  _Tot_Rows = -1 on ERROR of Loadind Rows

        if self._tTot_Rows < 0:
            return 'Error on loading Work sheet rows'
        if self._tTot_Rows <= 1:
            return 'No rows on the sheet'

        for nRow in range(1, self._tTot_Rows+1):
            self._Get_xlsx_Row(nRow)
            Des1Comp = Compact_Descr_String(self.Des1)
            Des2Comp = Compact_Descr_String(self.Des2)

            XLS_Row_List    = [int(nRow), self.Contab, self.Valuta,       # as it is in File
                              self.Des1,  self.Accr,   self.Addeb, self.Des2]
            XLS_Row_Compact = [int(nRow), self.Contab, self.Valuta,       # Desc1 & Desc2  compacted
                              Des1Comp,   self.Accr,   self.Addeb, Des2Comp]

            Checked_Row_List = self._Check_Values(XLS_Row_Compact)
            if not Checked_Row_List:
                self._tTot_NOK += 1
                self._tXlsx_Rows_NOK_List.append(XLS_Row_List)
            else:
                self._tTot_OK += 1
                self._tXLSX_Rows_Desc_Compact.append(Checked_Row_List)    # Descriptions compacted
                Data_Contab = Checked_Row_List[iRow_Contab]
                Data_Valuta = Checked_Row_List[iRow_Valuta]
                myRow = []
                Item = -1
                for Val in XLS_Row_List:
                    Item += 1
                    if Item == iRow_Contab:
                        myRow.append(Data_Contab)
                    elif Item == iRow_Valuta:
                        myRow.append(Data_Valuta)
                        iYear = int(Data_Valuta[0:4])
                        if iYear in self._tiYear_List or len(self._tiYear_List) >= 2:
                            pass
                        else:
                            self._tiYear_List.append(iYear)
                    else:
                        myRow.append(Val)
                self._tXLSX_Rows_From_Sheet.append(myRow)    # Descripritions as in Sheet, Date is str
        if self._tXlsx_Conto == FLASH or self._tXlsx_Conto == AMBRA or self._tXlsx_Conto == POSTA:
            self._Adjust_Rows_MostToLess()   # Invert order from Most Recent to Less
        return OK

    # ----------------------  Set tree rows list   ---------------------------------------  #
    def _Create_Xlsx_Lists(self):
        self._tiYear_List             = []
        self._tWihtout_Code_Tree_List = []
        self._tWith_Code_Tree_List    = []
        self._tTot_WithCode           = 0
        self._tTot_WithoutCode        = 0

        for Row in self._tXLSX_Rows_Desc_Compact:
            Row_Without_Code = []
            Row_With_Code    = []
            Full_Desc        = self.Descrip_Select(Row[iRow_Descr1], Row[iRow_Descr2])
            pass

            # Find for Full_Description of Row  a  String_To_Find on Codes Table
            # If more than one String_To_Find exists : Multiple match error

            Result = self._Find_StrToFind_InFullDesc(Row, Full_Desc)
            # return:      [NOK, []]  [OK, Found_List[0]]  [MULTI, 'multimatchText

            if Result[0] == NOK:
                self._tTot_WithoutCode += 1
                Row_Without_Code.append(Row[iRow_nRow])       # nRow
                Row_Without_Code.append(Row[iRow_Valuta])     # Valuta
                Amount = Row[iRow_Addeb]
                if type(Amount) is float:
                    if Amount == 0.0:
                        Amount = Row[iRow_Accr]
                else:
                    Amount = Row[iRow_Accr]
                Row_Without_Code.append(Amount)
                Row_Without_Code.append(Full_Desc)            # Full_Desc
                self._tWihtout_Code_Tree_List.append(Row_Without_Code)
            elif Result[0] == OK:
                Rec_Found = Result[1]
                self._tTot_WithCode += 1
                Row_With_Code.append(int(Row[iRow_nRow]))      # nRow
                Row_With_Code.append(Row[iRow_Contab])         # Contabile
                Row_With_Code.append(Row[iRow_Valuta])         # Valuta
                Row_With_Code.append(Rec_Found[iTR_TRdesc])    # TRdesc
                Row_With_Code.append(Row[iRow_Accr])           # Accred
                Row_With_Code.append(Row[iRow_Addeb])          # Addeb
                Row_With_Code.append(Rec_Found[iTR_TRcode])    # TRcode
                Row_With_Code.append(Rec_Found[iTR_TRfullDes]) # Full_Desc
                self._tWith_Code_Tree_List.append(Row_With_Code)
            else:
                return str(Result[1])    # Row matching  with multiple Str_ToFind
        return OK

    # --------------------------------------------------------------------------------------------
    # List_For_XLSX_Row_Control = [
    # [iRow_nRow,   INTEGER],
    # [iRow_Contab, DATE],    [iRow_Valuta, DATE],   [iRow_Descr1, STRING],
    # [iRow_Accr,   NUMERIC], [iRow_Addeb, NUMERIC], [iRow_Descr2, STRING]]
    def _Check_Values(self, XLS_Row_List):
        Xlsx_Row_List_Checked = []
        for Item_ToCheck in List_For_XLSX_Row_Control:
            Value = XLS_Row_List[Item_ToCheck[0]]
            Type = Item_ToCheck[1]
            ItemChecked = self._Check_Val(Value, Type)
            if ItemChecked is None:
                return []
            else:
                Xlsx_Row_List_Checked.append(ItemChecked)
        return Xlsx_Row_List_Checked  # as in Xlsx Rows
                                     # XLS_Row_List : nRow    Contab    Valuta    Descr1   Accred   Addeb   Descr2
    # -----------------------------------------------------------------------------------
    def _Check_Val(self, Item, Type):
        # List_For_XLSX_Row_Control = [
        # [iRow_nRow, INTEGER],
        # [iRow_Contab, DATE],  [iRow_Valuta, DATE],   [iRow_Descr1, STRING],
        # [iRow_Accr, NUMERIC], [iRow_Addeb, NUMERIC], [iRow_Descr2, STRING]  ]
        ItemType  = type(Item)
        if Type == STRING:            #   ---  String   -----  (Descriptions)
            if ItemType is str:
                return Item
            else:
                return ' '
        elif Type == INTEGER:         #   ---  Integer  -----   (Row Id number)
            if ItemType is int:
                return Item
            if ItemType is None:
                return 0
        elif Type == NUMERIC:         #   ---  Numeric  -----  (Accred  Addeb)
            if ItemType is float:
                if not Item:
                    return ' '
                return Item
            elif ItemType is int:
                if Item == 0:
                    return ' '
                return float(Item)
            # if ItemType is None:
            # return ' '
            return ' '
        elif Type == DATE:            #   ---  Date    -----
            Ckd_Date = self._Verify_Date(Item)
            if Ckd_Date:
                return Ckd_Date
            else:
                return None

    # -----------------------------------------------------------------------------------
    def _Get_xlsx_Row(self, nRow):
        if self._tXlsx_Conto == FIDEU:                                # Get columns for FIDEU
            self.Contab = self._Work_Sheet['A' + str(nRow)].value
            self.Valuta = self._Work_Sheet['B' + str(nRow)].value
            self.Des1   = self._Work_Sheet['C' + str(nRow)].value
            XlsxAccr    = self._Work_Sheet['D' + str(nRow)].value
            XlsxAddeb   = self._Work_Sheet['E' + str(nRow)].value
            self.Des2   = self._Work_Sheet['F' + str(nRow)].value
            # ----------   Credits and Debits  type are :  float   -------
            self.Accr  = Convert_To_Float(XlsxAccr)
            self.Addeb = Convert_To_Float(XlsxAddeb)

        elif self._tXlsx_Conto == FLASH or self._tXlsx_Conto == AMBRA:  # Get columns for Flash/AMBRA
            self.Contab = self._Work_Sheet['A' + str(nRow)].value
            self.Valuta = self._Work_Sheet['B' + str(nRow)].value
            self.Des1   = self._Work_Sheet['C' + str(nRow)].value
            XlsxAccr    = self._Work_Sheet['E' + str(nRow)].value
            XlsxAddeb   = self._Work_Sheet['G' + str(nRow)].value
            self.Des2   = ''
            # ----------   Credits and Debits  type are :  float   -------
            self.Accr  = Convert_To_Float(XlsxAccr)
            self.Addeb = -Convert_To_Float(XlsxAddeb)
            typeContab = type(self.Contab)
            typeValuta = type(self.Valuta)

            if typeContab is datetime and typeValuta is datetime:
                pass
            elif typeContab is datetime:
                self.Valuta = self.Contab
            elif typeValuta is datetime:
                self.Contab = self.Valuta
        # -----------------------------------------------------------------------------------
        #     A            B           C         D        E
        # DataContab.	DataVal.	Addebiti  Accred. Descrizione
        # -----------------------------------------------------------------------------------
        elif self._tXlsx_Conto == POSTA:                              # Get columns for POSTA
            self.Contab = self._Work_Sheet['A' + str(nRow)].value
            self.Valuta = self._Work_Sheet['B' + str(nRow)].value
            self.Des1   = ''  #self.Work_Sheet['C' + str(nRow)].value
            self.Accr   = self._Work_Sheet['D' + str(nRow)].value
            self.Addeb  = self._Work_Sheet['C' + str(nRow)].value
            self.Des2   = self._Work_Sheet['E' + str(nRow)].value
            if type(self.Addeb) is float or type(self.Addeb) is int:
                self.Addeb  = -self.Addeb
            typeContab = type(self.Contab)
            typeValuta = type(self.Valuta)
            if typeContab is datetime and typeValuta is datetime:
                pass
            elif typeContab is datetime:
                self.Valuta = self.Contab
            elif typeValuta is datetime:
                self.Contab = self.Valuta

        elif self._tXlsx_Conto == AMBRA:                              # Get columns for AMBRA
            pass

    # ----------------------------------------------------------------------------- #
    #  Data.contab   Data.val   Descriz.  Accred.val  Accred.  Addeb.val  Addeb.    #
    #       0            1        2           3          4         5        6       #
    # ----------------------------------------------------------------------------- #
    def _Adjust_Rows_MostToLess(self):
        Copy1 = self._Xlsx_Rows_From_Sheet.copy()
        self.XLSX_Rows_From_Sheet = []
        Copy2 = self._Xlsx_Rows_Desc_Compact.copy()
        self._Xlsx_Rows_Desc_Compact = []
        Index = self._Tot_OK -1
        for j in range(0, self._Tot_OK):
            Row1 = Copy1[Index]
            self.XLSX_Rows_From_Sheet.append(Row1)
            Row2 = Copy2[Index]
            self._Xlsx_Rows_Desc_Compact.append(Row2)
            Index -= 1

    # --------------------------------------------------------------------------------- #
    #  Workbook is the container of all Worksheets                                      #
    #  while the Worksheet is the container of Data of one Sheet                        #
    # --------------------------------------------------------------------------------- #
    def _Get_Work_Sheet_Rows(self, Filename):
        Work_Book = None
        try:
            Work_Book = load_workbook(Filename)
        except sqlite3.Error:
            self._tTot_Rows = -1
            return
        finally:
            self.SheetName   = Work_Book.sheetnames[0]   # always the first sheet
            self.Update_Selections(self.SheetName, Ix_Sheet_Name)
            self._Work_Sheet = Work_Book[self.SheetName]
            self._tTot_Rows  = self._Work_Sheet.max_row
            pass

    # -----------------------------------------------------------------------------------
    @classmethod
    def _Verify_Date(cls, DateToCheck):
        Type = type(DateToCheck)
        if Type is datetime:
            strDate = str(DateToCheck)
            myDate  = strDate[:10]
            return myDate
            # return DateToCheck
        DateTemplate = 'DD?MM?YYYY'
        if Type is not str:
            return []
        if len(DateToCheck) != 10:
            return []
        for i in range(0, 10):
            if DateTemplate[i] == '?':
                pass
            else:
                if not DateToCheck[i].isdecimal():
                    return []
        dateobject = datetime.strptime(DateToCheck, '%d/%m/%Y').date()
        return dateobject

    # -----------------------------------------------------------------------------------
    def Descrip_Select(self, Desc1, Desc2):
        self.Dummy = 1
        Len1 = 0
        Len2 = 0
        if type(Desc1) is str:
            Len1 = len(Desc1)
        if type(Desc2) is str:
            Len2 = len(Desc2)
        if not Len1 and not Len2:
            return '??????? Desciption_1 and 2  UNKNOWN  ?????????????'
        Full_Desc = Desc2
        if Len1 > Len2:
            Full_Desc = Desc1
        return Full_Desc

    # ------------------------------------------------------------------------------------
    def Check_The_Row(self, FullDesc, TRcode):
        for Desc in self.XLSX_Rows_From_Sheet:
            if Desc[iRow_nRow] == 96:
                pass
                for Rec in self._TR_Codes_Table:
                    if Rec[iTR_TRcode] == TRcode:
                        PRINT(FullDesc+ '   Record: ' + Rec)
                pass
# =======================================================================================
